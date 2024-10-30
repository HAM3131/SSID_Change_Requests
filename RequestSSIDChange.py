# RequestSSIDChange
# Purpose: Generate Excel sheet for SSID Change requests
# Author: Henry Manning
# Version: 0.0.8

import argparse
import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from pywintypes import com_error
from datetime import datetime
import win32com.client as win32
import sys
from SSIDErrors import SSIDError

excel = win32.gencache.EnsureDispatch('Excel.Application')

class SSID:
    def __init__(self, name, args):
        """Initialize SSID object

        param: name
            The name of the SSID to be modified
        param: args
            Namespace containing command line arguments
        """
        try:
            self.filename = name + '.xlsm'
            self.tmp_path = os.path.join('tmp', self.filename)
            self.name = name
            self.error_logging = args.error_logging
            self.verbose = args.verbose
            self.master_log_path = args.log_path
            self.log_path = os.path.join(os.path.dirname(args.log_path), f'{name}.log')
            self.summary = ''
            self.error_code = 0

            # Determine source path
            self.source_path = os.path.join(args.input_dir, self.filename)
            # Excel file with name of SSID isn't present in input directory. Check for dir
            if not os.path.isfile(self.source_path):
                ssid_folder = os.path.join(args.input_dir, name)
                max_mtime = 0
                if os.path.isdir(ssid_folder) and len(os.listdir(ssid_folder)) > 0:
                    for entry in os.listdir(os.path.join(args.input_dir, name)):
                        full_path = os.path.join(ssid_folder, entry)
                        mtime = os.stat(full_path).st_mtime
                        if mtime > max_mtime:
                            max_mtime = mtime
                            self.source_path = full_path
                else:
                    if os.path.isdir(os.path.join(args.input_dir, '! DELETED !', name)):
                        raise ValueError(f'No path found to source file for SSID `{name}`. SSID is in `! DELETED !` folder')
                    else:
                        raise LookupError(f'No path found to source file for SSID `{name}`. dir `{ssid_folder}` exists: {os.path.isdir(ssid_folder)}')
            
            if self.source_path.endswith('.pdf'):
                raise NameError(f'No path found to source file for SSID `{name}`. File in directory is a pdf: `{self.source_path}`')

            # Determine output to finally save file
            if args.file_input and args.output is not None:
                if not os.path.exists(args.output):
                    os.makedirs(args.output)
                self.output_path = args.output
            elif args.output is not None:
                self.output_path = args.output + '.xlsm'
            else:
                self.output_path = '.'

            # Create `tmp` directory if not present 
            if not os.path.isdir('tmp'):
                os.makedirs('tmp')

            # Copy file into tmp folder for working
            if not self.source_path.endswith('.xlsm'):
                copy_excel_as_xlsm(self.source_path, os.path.join(os.getcwd(), self.tmp_path))
            else:
                with open(self.source_path, 'rb') as f:
                    contents = f.read()
                with open(self.tmp_path, 'wb') as f:
                    f.write(contents)

        except ValueError as e:
            self.log_error(f'ERROR: SSID.__init__(`{name}`, args): {e}', SSIDError.SSID_DELETED)

        except NameError as e:
            self.log_error(f'ERROR: SSID.__init__(`{name}`, args): {e}', SSIDError.SSID_IS_PDF)

        except LookupError as e:
            self.log_error(f'ERROR: SSID.__init__(`{name}`, args): {e}', SSIDError.SSID_ERROR)

        except com_error as e:
            self.log_error(f'ERROR: SSID.__init__(`{name}`, args): {e}', SSIDError.SSID_CANT_OPEN_SPREADSHEET)
        
        else:
            self.log(f'SSID `{self.name}` initialized successfully')

    def log(self, message):
        message = message + f' [{datetime.now().strftime("%H:%M:%S")}]'
        with open(self.master_log_path, 'a') as f:
            f.write(message + '\n')
        with open(self.log_path, 'a') as f:
            f.write(message + '\n')
        if self.verbose:
            print(message)
        
    def log_error(self, message, error_code):
        self.error_code = error_code
        message = message + f' [{datetime.now().strftime("%H:%M:%S")}]'
        if os.path.isfile(self.tmp_path):
            os.remove(self.tmp_path)
        with open(self.master_log_path, 'a') as f:
            f.write(message + '\n')
        with open(self.log_path, 'a') as f:
            f.write(message + '\n')
        if self.error_logging:
            print(message)
        
    def change_primary_manager(self, args):
        """Make appropriate updates to the spreadsheet for a primary manager change

        param: args
            string with previous and new manager separated by semicolon: `<Previous Manager>;<New Manager>`

        return: None 
        """
        try:
            # Alias variables
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            old_manager, new_manager = args.split(';')

            # Modify `Acct Info` sheet
            ws = wb['Acct Info']
            if not old_manager.lower() == ws['B30'].value.lower():
                raise ValueError(f'previous primary manager = `{ws["B30"].value}`, expected `{old_manager}`')
            ws['B28'] = 'Yes'
            summary_previous_manager = ws['B30'].value
            ws['B30'] = new_manager
            
            wb.save(self.tmp_path)
            wb.close()
            self.summary += f'Change primary manager to {new_manager} - previous manager was {summary_previous_manager}. '
        
        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_primary_manager(): {e}', SSIDError.SSID_ERROR)

        except KeyError as e:
           self.log_error(f'ERROR: `{self.name}` - SSID.change_primary_manager(): {e}', SSIDError.SSID_MISSING_SPREADSHEET_SHEET)
             
        else:
            self.log(f'Primary manager changed from `{summary_previous_manager}` to `{new_manager}` for SSID `{self.name}`')
    
    def change_secondary_manager(self, args):
        """Make appropriate updates to the spreadsheet for a secondary manager change

        param: args
            string with previous and new managers separated by semicolon: `<Previous Manager>;<New Manager>`

        return: None 
        """
        try:
            # Alias variables
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            old_manager, new_manager = args.split(';')

            # Modify `Acct Info` sheet
            ws = wb['Acct Info']
            if not old_manager.lower() == ws['B32'].value.lower():
                raise ValueError(f'previous secondary manager = `{ws["B32"].value}`, expected `{old_manager}`')
            ws['B28'] = 'Yes'
            summary_previous_manager = ws['B32'].value
            ws['B32'] = new_manager
            
            wb.save(self.tmp_path)
            wb.close()
            self.summary += f'Change secondary manager to {new_manager} - previous manager was {summary_previous_manager}. '
        
        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_secondary_manager(): {e}', SSIDError.SSID_ERROR)
            return False
    
        except KeyError as e:
           self.log_error(f'ERROR: `{self.name}` - SSID.change_secondary_manager(): {e}', SSIDError.SSID_MISSING_SPREADSHEET_SHEET)

        else:
            self.log(f'Secondary manager changed from `{old_manager}` to `{new_manager}` for SSID `{self.name}`')
            return True

    def change_manager(self, args):
        """Make appropriate updates to the spreadsheet for a primary manager change

        param: args
            Namespace with `change_manager` defined

        return: None 
        """
        try:
            # Alias variables
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            old_manager, _ = args.change_manager.split(';')

            ws = wb['Acct Info']
            primary_manager = ws['B30'].value
            if primary_manager is None:
                primary_manager = ''
            secondary_manager = ws['B32'].value
            if secondary_manager is None:
                secondary_manager = ''
            if primary_manager.lower() == old_manager.lower():
                self.log(f'change_manager() selected `primary manager` for SSID `{self.name}`')
                self.change_primary_manager(args.change_manager)
            elif secondary_manager.lower() == old_manager.lower():
                self.log(f'change_manager() selected `secondary manager` for SSID `{self.name}`')
                self.change_secondary_manager(args.change_manager)
            else:
                raise ValueError(f'Neither primary ({primary_manager}) nor secondary ({secondary_manager}) manager matches expected previous manager: `{old_manager}`')
        
        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_manager(): {e}', SSIDError.INVALID_PREVIOUS_MANAGER)
        
        except:
            self.log_error(f'ERROR: `{self.name}` - unknown error in `change_manager()`', SSIDError.SSID_ERROR)

    def change_primary_account_custodian(self, args):
        try:
            # Alias variables
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            old_custodian, new_custodian = args.split(';')

            ws = wb['Acct Info']
            row = find_row(ws, 'A', 'Primary Acct Custodian')
            if row is None:
                raise TypeError('A row matching `Primary Acct Custodian` could not be found.')

            current_custodian = ws[f'B{row}'].value

            if not current_custodian == old_custodian and not old_custodian == 'any':
                raise ValueError(f'previous primary account custodian = `{current_custodian}`, expected `{old_custodian}`')
            

            if current_custodian == new_custodian:
                self.log(f'Primary Custodian `{new_custodian}` already set for SSID {self.name}. No change.')
                return

            ws[f'B{row}'] = new_custodian
            ws['B28'] = 'Yes'

            wb.save(self.tmp_path)
            wb.close()
            self.summary += f'Change secondary manager to {new_custodian} - previous manager was {current_custodian}. '

        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_primary_account_custodian(): {e}', SSIDError.SSID_ERROR)
            return False
    
        except KeyError as e:
           self.log_error(f'ERROR: `{self.name}` - SSID.change_primary_account_custodian(): {e}', SSIDError.SSID_MISSING_SPREADSHEET_SHEET)

        except TypeError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_primary_account_custodian(): {e}', SSIDError.MISSING_ROW)

        except:
            self.log_error(f'ERROR: `{self.name}` - unknown error in `change_manager()`', SSIDError.SSID_ERROR)

        else:
            self.log(f'Primary account custodian changed from `{old_custodian}` to `{new_custodian}` for SSID `{self.name}`')
            return True
    
    def change_authorized_users(self, args):
        try:
            # Alias variables
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            old_user, new_user = args.split(';')

            ws = wb['Acct Info']
            row = int(find_row(ws, 'A', 'Authorized Users'))
            if row is None:
                raise TypeError('A row matching `Authorized Users` could not be found.')

            empty_rows = 0
            while ws[f'A{row + empty_rows+1}'].value is None:
                empty_rows += 1
            current_users = []
            for i in range(empty_rows+1):
                current_user = ws[f'B{row + i}'].value
                if current_user is not None:
                    current_users.append(current_user)

            if not old_user in current_users and not old_user == 'any':
                raise ValueError(f'previous authorized users = `{current_users}`, expected `{old_user}` to exist')
            
            if new_user in current_users:
                self.log(f'Authorized user `{new_user}` already present for SSID {self.name}. No change.')
                wb.close()
                return
            
            for i in range(empty_rows+1):
                ws[f'B{row + i}'] = ''

            ws[f'B{row}'] = new_user
            ws['B28'] = 'Yes'

            wb.save(self.tmp_path)
            wb.close()
            self.summary += f'Change authorized user to {new_user} - previous users were {current_users}. '

        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_authorized_users(): {e}', SSIDError.SSID_ERROR)
            return False
    
        except KeyError as e:
           self.log_error(f'ERROR: `{self.name}` - SSID.change_authorized_users(): {e}', SSIDError.SSID_MISSING_SPREADSHEET_SHEET)

        except TypeError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.change_authorized_users(): {e}', SSIDError.MISSING_ROW)

        except:
            self.log_error(f'ERROR: `{self.name}` - unknown error in `change_manager()`', SSIDError.SSID_ERROR)

        else:
            self.log(f'Authorized user changed from `{old_user}` to `{new_user}` for SSID `{self.name}`')
            return True

    def remove_legacy_drawings(self):
        """Remove the broken legacy drawings on sheets `DB2 UNIX`, `Mainframe`, and `Other`
        """
        try:
            # Load the workbook in `tmp` directory
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            
            broken_drawings = ['DB2 UNIX', 'DB2 AIX', 'Mainframe', 'Other']
            # Remove legacy drawings from broken pages
            for sheet in list(set(broken_drawings) & set([sheet.title for sheet in wb._sheets])):
                wb[sheet].legacy_drawing = None
            
            # Save workbook back to `tmp` folder
            wb.save(self.tmp_path)
            wb.close()

        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.remove_legacy_drawings(): {e}', SSIDError.SSID_ERROR)
        
        else:
            self.log(f'Legacy drawings removed from SSID `{self.name}` successfully')

    def write_summary(self):
        """Write the summary of all actions taken onto the `Summary` page
        """
        try:
            if self.summary == '':
                raise ValueError('Summary is empty, no changes have been made. Cannot create summary')

            # Modify `Summary` sheet
            wb = load_workbook(self.tmp_path, read_only=False, keep_vba=True)
            ws = wb['Summary']
            merged_cells_range = ws.merged_cells.ranges
            for merged_cell in merged_cells_range:
                _, top, _, _ = merged_cell.bounds
                if top >= 12:
                    merged_cell.shift(0,7)
            ws.insert_rows(12,7)
            cellrange = ws['C12:L17']
            for row in cellrange:
                for cell in row:
                    border = copy(cell.border)
                    if cell.row == 12:
                        border.top = Side(style='medium')
                    elif cell.row == 17:
                        border.bottom = Side(style='medium')
                    if cell.column == 3:
                        border.left = Side(style='medium')
                    elif cell.column == 12:
                        border.right = Side(style='medium')
                    cell.border = border
            ws.merge_cells(start_row=12, start_column=3, end_row=17, end_column=12)
            cell = ws['C12']
            cell.alignment = Alignment(vertical='top')

            ws['A12'] = 'Date'
            ws['A13'] = datetime.today().strftime('%m/%d/%Y')
            ws['B12'] = 'REQ'
            ws['C12'] = self.summary

            # Save workbook
            wb.save(self.tmp_path)
            wb.close()

        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.write_summary(): {e}', SSIDError.NO_CHANGES)

        else:
            self.log(f'Summary written for SSID `{self.name}` successfully')

    def output(self):
        """Save the spreadsheet to the final output destination
        """
        try:
            if '.xlsm' not in self.output_path:
                self.output_path = os.path.join(self.output_path, f'{datetime.now().strftime("%Y-%m-%d_%H%M")}_{self.name}.xlsm')

            if self.error_code == 0:
                with open(self.tmp_path, 'rb') as f:
                    contents = f.read()
                with open(self.output_path, 'wb') as f:
                    f.write(contents)
                print(f'SSID `{self.name}` completed and output at {self.output_path}')
                os.remove(self.tmp_path)
            else:
                print(f'SSID `{self.name}` errored during process - cannot save')
            if os.path.isdir('tmp') and len(os.listdir('tmp')) == 0:
                os.removedirs('tmp')

        except ValueError as e:
            self.log_error(f'ERROR: `{self.name}` - SSID.output(): {e}', SSIDError.SSID_ERROR)
        
        except WindowsError as e:
            log(self.master_log_path, f'Failed to delete `tmp` dir: {e}')

        else:
            self.log(f'SSID `{self.name}` output successfully')

def find_row(sheet, column, search_string):
    """Find the row in the given column of the sheet which contains a specific string

    param: sheet
        the spreadsheet to search in
    param: column
        the column to search in
    param: search_string
        the precise string to look for
    
    return:
        the row number in string format if found, or None if not found
    """
    for cell in sheet[column]:
        if cell.value == search_string:
            return cell.row
    return None

def copy_excel_as_xlsm(source_path, output_path):
     try:
        if os.path.isfile(output_path):
            os.remove(output_path)
        wb = excel.Workbooks.Open(source_path)
        wb.SaveAs(output_path, 52)
        wb.Close(SaveChanges=False)
     except IOError:
        print('Error')

def parse_args():
    """Define an argparse parser and return the parsed arguments

    return: parsed_args
        Namespace containing parsed arguments
    """
    parser = argparse.ArgumentParser(
                    prog='RequestSSIDChange',
                    description='Generates an Excel spreadsheet for an SSID Change request',
                    epilog='Contact Henry Manning for suggestions. [henry_manning@cinfin.com]')
    
    parser.add_argument('filename',
                        type=str,
                        help='Name of the SSID being managed')

    parser.add_argument('-cau',
                        '--change-authorized-users',
                        type=str,
                        help='Name of previou and new authorized user. Expected format: `<Previous User>;<New User>`')

    parser.add_argument('-cm',
                        '--change-manager',
                        type=str,
                        help='Name of previous and new manager (either primary or secondary). Expected format: `<Previous Manager>;<New Manager>`')

    parser.add_argument('-cpac',
                        '--change-primary-account-custodian',
                        type=str,
                        help='Name of previous and new primary account custodian.  Expected format: `<Previous Custodian>;<New Custodian>`')

    parser.add_argument('-cpm',
                        '--change-primary-manager',
                        type=str,
                        help='Name of previous and new primary manager. Expected format: `<Previous Manager>;<New Manager>`')
    
    parser.add_argument('-csm',
                        '--change-secondary-manager',
                        type=str,
                        help='Name of previous and new secondary manager. Expected format: `<Previous Manager>;<New Manager>`')
    
    parser.add_argument('-e',
                        '--error-logging',
                        action='store_true',
                        help='Flag to turn on error logging')
    
    parser.add_argument('-f',
                        '--file-input',
                        action='store_true',
                        help='Specify a text file with a different SSID on each line instead of a single SSID to change')

    parser.add_argument('-i',
                        '--input-dir',
                        type=str,
                        default='\\\\wfshq1\\acna\\SSID Forms\\New SSID Forms',
                        help='Path to dir which SSID spreadsheets are in. Defaults to `\\\\wfshq1\\acna\\SSID Forms\\New SSID Forms`')
    
    parser.add_argument('-o',
                        '--output',
                        type=str,
                        default=None,
                        help='Filename for output, or directory to output to if using a file as input')
    
    parser.add_argument('-v',
                        '--verbose',
                        action='store_true',
                        default=False,
                        help='Flag to turn on verbose output')

    parsed_args = parser.parse_args()

    # Ensure `-cpm, -csm, -cm` flags are not used together
    if len([flag for flag in [parsed_args.change_primary_manager, parsed_args.change_secondary_manager, parsed_args.change_manager] if flag is not None]) > 1:
        raise ValueError(f'INPUT ERROR: Multiple flags for changing managers used. Use just one of `-cpm, --change-primary-manager`, `-csm, --change-secondary-manager`, and `-cm, --change-manager`')

    # Ensure `-cpm, --change-primary-manager` value has proper syntax, including a semicolon separating new and old managers
    if parsed_args.change_primary_manager is not None and not ';' in parsed_args.change_primary_manager:
        raise ValueError(f'`INPUT ERROR: -cpm, --change-primary-manager` flag used without proper syntax: `{parsed_args.change_primary_manager}`\n'
                        +f'\tExpected: `<Previous Manager>;<New Manager>`')
    
    # Ensure `-csm, --change-secondary-manager` value has proper syntax, including a semicolon separating new and old managers
    if parsed_args.change_secondary_manager is not None and not ';' in parsed_args.change_secondary_manager:
        raise ValueError(f'`INPUT ERROR: -csm, --change-secondary-manager` flag used without proper syntax: `{parsed_args.change_secondary_manager}`\n'
                        +f'\tExpected: `<Previous Manager>;<New Manager>`')
    
    # Ensure `-cm, --change-manager` value has proper syntax, including a semicolon separating new and old managers
    if parsed_args.change_manager is not None and not ';' in parsed_args.change_manager:
        raise ValueError(f'INPUT ERROR: `-cm, --change-manager` flag used without proper syntax: `{parsed_args.change_manager}`\n'
                        +f'\tExpected: `<Previous Manager>;<New Manager>`')

    return parsed_args

def get_ssid_list(args):
    """Generate list of SSID objects from input

    param: args
        Namespace including `file_input` and `filename`
    
    return: SSIDs
        list of SSID objects
    """
    if args.file_input:
        with open(args.filename) as f:
            ssid_names = [line.strip() for line in f.readlines()]
        SSIDs = [SSID(name, args) for name in ssid_names]
    else:
        SSIDs = [SSID(args.filename, args)]
    
    return SSIDs

def log(path, message):
    with open(path, 'a') as f:
        f.write(message + f' [{datetime.now().strftime("%H:%M:%S")}]\n')
    print(message)

def execute_changes(args):
    """Initialize list of SSIDs, make appropriate edits to spreadsheets, and save

    param: args
        Namespace including `change_primary_manager` and other key details of command line arguments
    """
    log(args.log_path, '\033[1;34m***Initializing SSID objects***\033[22;0m')
    SSIDs = get_ssid_list(args)

    if args.change_primary_manager is not None:
        log(args.log_path, '\033[1;34m***Changing primary managers***\033[22;0m')
        [ssid.change_primary_manager(args.change_primary_manager) for ssid in SSIDs if ssid.error_code == 0]

    if args.change_secondary_manager is not None:
        log(args.log_path, '\033[1;34m***Changing secondary managers***\033[22;0m')
        [ssid.change_secondary_manager(args.change_primary_manager) for ssid in SSIDs if ssid.error_code == 0]

    if args.change_manager is not None:
        log(args.log_path, '\033[1;34m***Changing managers***\033[22;0m')
        [ssid.change_manager(args) for ssid in SSIDs if ssid.error_code == 0]
    
    if args.change_primary_account_custodian is not None:
        log(args.log_path, '\033[1;34m***Changing primary account custodians***\033[22;0m')
        [ssid.change_primary_account_custodian(args.change_primary_account_custodian) for ssid in SSIDs if ssid.error_code == 0]
    
    if args.change_authorized_users is not None:
        log(args.log_path, '\033[1;34m***Changing authorized users***\033[22;0m')
        [ssid.change_authorized_users(args.change_authorized_users) for ssid in SSIDs if ssid.error_code == 0]
    
    log(args.log_path, '\033[1;34m***Writing change summaries***\033[22;0m')
    [ssid.write_summary() for ssid in SSIDs if ssid.error_code == 0]

    # Remove legacy drawings
    log(args.log_path, '\033[1;34m***Removing legacy drawings***\033[22;0m')
    [ssid.remove_legacy_drawings() for ssid in SSIDs if ssid.error_code == 0]
    
    log(args.log_path, '\033[1;34m***Saving successfully modified files***\033[22;0m')
    [ssid.output() for ssid in SSIDs if ssid.error_code == 0]

    # Calculate totals for different error codes
    successful_edits = len([ssid for ssid in SSIDs if ssid.error_code == 0])
    failed_edits = len([ssid for ssid in SSIDs if (ssid.error_code != 0 and ssid.error_code != 1 and ssid.error_code != SSIDError.SSID_DELETED)])
    deleted_ssids = len([ssid for ssid in SSIDs if ssid.error_code == SSIDError.SSID_DELETED])

    # Print meta results
    log(args.log_path, f'\033[1;32mFile editing completed\033[22;0m - {successful_edits}/{successful_edits + failed_edits} edited successfully')
    log(args.log_path, f'\033[1;31mSome SSIDs have been deleted in the past\033[22;0m - {deleted_ssids}')
    
    # Write list of successfully edited files
    with open(os.path.join(os.path.dirname(args.log_path), '.edits_successful'), 'w') as f:
        for ssid in [ssid for ssid in SSIDs if ssid.error_code == 0]:
            f.write(ssid.name + '\n')
    
    # Write list of failed to edit files, or previously deleted files
    with open(os.path.join(os.path.dirname(args.log_path), '.edits_failed'), 'w') as f:
        for error_code in SSIDError:
            ssids_with_error = [ssid for ssid in SSIDs if ssid.error_code == error_code]
            if len(ssids_with_error) > 0:
                f.write('[' + str(error_code) + ']\n')
                for ssid in ssids_with_error:
                    f.write(ssid.name + '\n')
                f.write('\n')

def create_log_file():
    log_path = os.path.join('logs', datetime.now().strftime('%Y-%m-%d_%H%M%S'))
    if not os.path.isdir(log_path):
        os.makedirs(log_path)
    log_path = os.path.join(log_path, '.SSID_Changes.log')
    with open(log_path, 'w') as f:
        f.write(f'[{datetime.now().strftime("%Y-%m-%d_%H:%M:%S")}] Log Created\n')
        f.write(f'ARGS: {sys.argv}')
    return log_path

def main():
    """Generate new excel sheet
    
    return: status
        0 if success, 1 if any errors are encountered
    """
    status = 0

    try:
        args = parse_args()
        args.log_path = create_log_file()
        execute_changes(args)
    except ValueError as e:
        print(e)

    return status

if __name__ == '__main__':
    main()

excel.Quit()